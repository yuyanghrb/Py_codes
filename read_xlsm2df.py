from openpyxl import load_workbook


def cell2df(_cell_range):
    '''
    This function will read a range of cell in xlsm worksheet
    Transform to a panda DataFrame
    "demand_instore =  cell2df(['N14','T45'])"
    '''
    _start_cell = _cell_range[0]
    _end_cell = _cell_range[1]  
    _data_rows = []
    for _row in ws[_start_cell:_end_cell]:
        _data_cols = []
        for _cell in _row:
            if _cell.value is None:
                _data_cols.append(0)
            
            else: 
                _data_cols.append(_cell.value)
        _data_rows.append(_data_cols)
    _colNames = []
    for i in range(1,(1+len(_data_cols))):
        _colNames.append((i))
    _rowNames = []
    for i in range(1,(1+len(_data_rows))):
         _rowNames.append(i)
    _df = pd.DataFrame(_data_rows, index = _rowNames, columns = _colNames)
    return _df

wb = load_workbook('Test.xlsm', data_only=True, keep_vba=True)
ws = wb['Tabname']

Test_DF = cell2df(['A1', 'D5'])
    